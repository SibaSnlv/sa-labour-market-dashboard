import pandas as pd
import numpy as np
import openpyxl
import json
import os
import re
from sklearn.linear_model import LinearRegression
from scipy import stats

QLFS_FILE = "QLFS Trends 2008-2026Q1.xlsx"
GDP_FILE  = "Provincial GDP tables 2013 - 2024.xlsx"
GEOJSON_FILE = "south_africa_provinces.geojson.json"
CLEAN_DIR = "clean_data/"
os.makedirs(CLEAN_DIR, exist_ok=True)

PROVINCE_FIX = {
    "KwaZulu Natal": "KwaZulu-Natal",
}

OFFICIAL_PROVINCES = [
    "Western Cape", "Eastern Cape", "Northern Cape", "Free State",
    "KwaZulu-Natal", "North West", "Gauteng", "Mpumalanga", "Limpopo"
]

QUARTER_MONTH_MAP = {
    "Jan-Mar": 1, "Apr-Jun": 4, "Jul-Sep": 7, "Oct-Dec": 10
}

def parse_qlfs_date(label):
    if label is None:
        return None
    match = re.match(
        r"(Jan-Mar|Apr-Jun|Jul-Sep|Oct-Dec) (\d{4})", str(label).strip()
    )
    if not match:
        return None
    period, year = match.groups()
    return pd.Timestamp(year=int(year), month=QUARTER_MONTH_MAP[period], day=1)

def get_qlfs_quarter_columns(ws, header_row=2, start_col=2, end_col=74):
    quarter_cols = {}
    for col in range(start_col, end_col + 1):
        date = parse_qlfs_date(ws.cell(row=header_row, column=col).value)
        if date is not None:
            quarter_cols[col] = date
    return quarter_cols

PROVINCE_BLOCK_ROWS = {
    "South Africa": 5,
    "Western Cape": 21,
    "Eastern Cape": 69,
    "Northern Cape": 133,
    "Free State": 149,
    "KwaZulu Natal": 205,
    "North West": 253,
    "Gauteng": 269,
    "Mpumalanga": 349,
    "Limpopo": 365,
}

METRIC_OFFSETS = {
    "working_age_population": 1,
    "labour_force": 2,
    "employed": 3,
    "unemployed": 4,
    "outside_labour_force": 5,
    "participation_rate": 7,
    "absorption_rate": 8,
    "inactivity_rate": 9,
    "unemployment_rate": 11,
}

def parse_table_2_3(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table 2.3"]
    quarter_cols = get_qlfs_quarter_columns(ws)
    records = []
    for province_raw, start_row in PROVINCE_BLOCK_ROWS.items():
        if province_raw == "South Africa":
            continue
        province = PROVINCE_FIX.get(province_raw, province_raw)
        for col, date in quarter_cols.items():
            record = {"province": province, "date": date}
            for metric_name, offset in METRIC_OFFSETS.items():
                record[metric_name] = ws.cell(row=start_row + offset, column=col).value
            records.append(record)
    return pd.DataFrame(records)

AGE_BLOCK_ROWS = {
    "15-64": 5, "15-24": 21, "25-34": 37,
    "35-44": 53, "45-54": 69, "55-64": 85,
}

AGE_METRIC_OFFSETS = {
    "population": 1, "labour_force": 2, "employed": 3,
    "unemployed": 4, "outside_labour_force": 5,
    "participation_rate": 7, "absorption_rate": 8,
    "inactivity_rate": 9, "unemployment_rate": 11,
}

def parse_table_2_2(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table2.2"]
    quarter_cols = get_qlfs_quarter_columns(ws)
    records = []
    for age_group, start_row in AGE_BLOCK_ROWS.items():
        for col, date in quarter_cols.items():
            record = {"age_group": age_group, "date": date}
            for metric_name, offset in AGE_METRIC_OFFSETS.items():
                record[metric_name] = ws.cell(row=start_row + offset, column=col).value
            records.append(record)
    return pd.DataFrame(records)

def build_youth_15_34(df_age):
    youth = df_age[df_age["age_group"].isin(["15-24", "25-34"])].copy()
    youth_agg = youth.groupby("date").agg({
        "population": "sum", "labour_force": "sum",
        "employed": "sum", "unemployed": "sum",
    }).reset_index()
    youth_agg["unemployment_rate"] = youth_agg["unemployed"] / youth_agg["labour_force"]
    youth_agg["age_group"] = "15-34"
    return youth_agg

INDUSTRY_BLOCK_ROWS = {
    "Agriculture": 5, "Mining": 16, "Manufacturing": 27,
    "Utilities": 38, "Construction": 49, "Trade": 66,
    "Transport": 77, "Finance": 88,
    "Community and social services": 99, "Private households": 110,
}

PROVINCE_ORDER_IN_INDUSTRY_TABLE = [
    "Western Cape", "Eastern Cape", "Northern Cape", "Free State",
    "KwaZulu Natal", "North West", "Gauteng", "Mpumalanga", "Limpopo"
]

def parse_table_3_2(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table3.2"]
    quarter_cols = get_qlfs_quarter_columns(ws)
    records = []
    for sector, sector_row in INDUSTRY_BLOCK_ROWS.items():
        for i, province_raw in enumerate(PROVINCE_ORDER_IN_INDUSTRY_TABLE):
            province = PROVINCE_FIX.get(province_raw, province_raw)
            row = sector_row + 1 + i
            for col, date in quarter_cols.items():
                records.append({
                    "sector": sector.strip(),
                    "province": province,
                    "date": date,
                    "employed": ws.cell(row=row, column=col).value
                })
    df = pd.DataFrame(records)
    df["employed"] = df["employed"].replace(".", None)
    return df

GDP_SECTION_HEADER_ROWS = {
    "current_price_rand": 3,
    "current_price_pct": 16,
    "constant_price_rand": 29,
    "constant_price_growth_pct": 42,
}

def parse_table_21(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table 21"]
    all_records = []
    for section_name, header_row in GDP_SECTION_HEADER_ROWS.items():
        year_cols = {}
        for col in range(2, 14):
            year_val = ws.cell(row=header_row, column=col).value
            if year_val is not None:
                year_cols[col] = int(year_val)
        for i, province_raw in enumerate(OFFICIAL_PROVINCES):
            row = header_row + 1 + i
            for col, year in year_cols.items():
                all_records.append({
                    "province": province_raw,
                    "year": year,
                    "metric": section_name,
                    "value": ws.cell(row=row, column=col).value
                })
    df_long = pd.DataFrame(all_records)
    df_wide = df_long.pivot_table(
        index=["province", "year"], columns="metric", values="value"
    ).reset_index()
    return df_wide

def load_geojson_for_powerbi(filepath):
    with open(filepath, "r") as f:
        geo = json.load(f)
    geo_provinces = sorted([f["properties"]["name"] for f in geo["features"]])
    expected = sorted(OFFICIAL_PROVINCES)
    if geo_provinces == expected:
        print("GeoJSON province names match exactly. No remapping needed.")
    else:
        print("MISMATCH:", set(geo_provinces) ^ set(expected))
    return geo

def calculate_derived_metrics(df):
    df = df.copy()
    df["year"] = df["date"].dt.year
    df["quarter"] = df["date"].dt.quarter
    df["quarter_label"] = "Q" + df["quarter"].astype(str) + " " + df["year"].astype(str)
    df = df.sort_values(["province", "date"])
    df["unemployment_rate_qoq"] = df.groupby("province")["unemployment_rate"].diff()
    df["unemployment_rate_yoy"] = df.groupby("province")["unemployment_rate"].diff(4)
    return df

def aggregate_to_annual(df_labour):
    return df_labour.groupby(["province", "year"]).agg({
        "unemployment_rate": "mean", "absorption_rate": "mean",
        "employed": "mean", "labour_force": "mean",
        "working_age_population": "mean",
    }).reset_index()

def parse_gender_unemployment(filepath):
    """Women = row 50, Men = row 75 (LU1 unemployment rate)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table 2"]
    quarter_cols = get_qlfs_quarter_columns(ws)
    GENDER_ROWS = {"Women": 50, "Men": 75}
    records = []
    for gender, row in GENDER_ROWS.items():
        for col, date in quarter_cols.items():
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip() != "-" and str(value).strip() != " -":
                try:
                    records.append({
                        "date": date,
                        "gender": gender,
                        "unemployment_rate": float(str(value).replace(" ", "")),
                        "year": date.year,
                        "quarter": date.quarter,
                        "quarter_label": f"Q{date.quarter} {date.year}",
                    })
                except ValueError:
                    pass
    return pd.DataFrame(records)

NEET_ROW_MAP = {
    "Western Cape": 16, "Eastern Cape": 17, "Northern Cape": 18,
    "Free State": 19, "KwaZulu Natal": 20, "North West": 21,
    "Gauteng": 22, "Mpumalanga": 23, "Limpopo": 24,
}

def parse_neet_by_province(filepath):
    """NEET youth 15-34 by province. Available from Q3 2012 onward."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table7c"]
    quarter_cols = get_qlfs_quarter_columns(ws)
    records = []
    for province_raw, row in NEET_ROW_MAP.items():
        province = PROVINCE_FIX.get(province_raw, province_raw)
        for col, date in quarter_cols.items():
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip() not in ("-", " -", ""):
                try:
                    records.append({
                        "province": province,
                        "date": date,
                        "year": date.year,
                        "quarter": date.quarter,
                        "quarter_label": f"Q{date.quarter} {date.year}",
                        "neet_thousands": float(str(value).replace(" ", "")),
                    })
                except ValueError:
                    pass
    return pd.DataFrame(records)

NATIONAL_SECTOR_ROWS = {
    "Agriculture": 6, "Mining": 7, "Manufacturing": 8,
    "Utilities": 9, "Construction": 10, "Trade": 11,
    "Transport": 12, "Finance": 13,
    "Community and social services": 14, "Private households": 15,
}

def parse_national_sector_totals(filepath):
    """National employment by sector (Both sexes) from Table 3.1."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table3.1"]
    quarter_cols = get_qlfs_quarter_columns(ws)
    records = []
    for sector, row in NATIONAL_SECTOR_ROWS.items():
        for col, date in quarter_cols.items():
            value = ws.cell(row=row, column=col).value
            if value is not None:
                try:
                    records.append({
                        "sector": sector,
                        "date": date,
                        "year": date.year,
                        "quarter": date.quarter,
                        "quarter_label": f"Q{date.quarter} {date.year}",
                        "employed_thousands": float(value),
                    })
                except (ValueError, TypeError):
                    pass
    df = pd.DataFrame(records)
    df = df.sort_values(["sector", "date"])
    df["employed_yoy_growth"] = (
        df.groupby("sector")["employed_thousands"].pct_change(4) * 100
    )
    total_by_quarter = df.groupby("date")["employed_thousands"].transform("sum")
    df["sector_share_pct"] = (df["employed_thousands"] / total_by_quarter) * 100
    latest = df[df["date"] == df["date"].max()][["sector", "employed_thousands"]].copy()
    latest["rank_by_employment"] = (
        latest["employed_thousands"].rank(ascending=False).astype(int)
    )
    df = df.merge(latest[["sector", "rank_by_employment"]], on="sector", how="left")
    return df

def build_projections(df_labour, projection_quarters=8):
    """
    Linear trend projection per province, 8 quarters forward.
    Three scenarios: Baseline, Optimistic, Pessimistic.
    Includes 95% confidence interval bounds.
    Uses data from 2015 onward for cleaner trend fitting.
    """
    df_model = df_labour[df_labour["year"] >= 2015].copy()
    projection_records = []
    summary_records = []

    for province in OFFICIAL_PROVINCES:
        prov = df_model[df_model["province"] == province].sort_values("date").copy()
        if len(prov) < 8:
            continue

        prov["date_num"] = (prov["date"] - prov["date"].min()).dt.days
        X = prov["date_num"].values.reshape(-1, 1)
        y = prov["unemployment_rate"].values

        model = LinearRegression()
        model.fit(X, y)
        r_squared = model.score(X, y)
        slope = model.coef_[0]
        slope_per_quarter = slope * 91

        last_date    = prov["date"].max()
        last_numeric = prov["date_num"].max()
        last_rate    = prov["unemployment_rate"].iloc[-1]

        residuals = y - model.predict(X)
        n = len(y)
        x_mean = prov["date_num"].mean()
        s = np.sqrt(np.sum(residuals**2) / (n - 2))

        for i in range(1, projection_quarters + 1):
            future_date = last_date + pd.DateOffset(months=3 * i)
            future_num  = last_numeric + (91 * i)
            projected   = float(np.clip(model.predict([[future_num]])[0], 0, 100))

            se = s * np.sqrt(
                1 + 1/n + (future_num - x_mean)**2 /
                np.sum((prov["date_num"] - x_mean)**2)
            )
            margin = stats.t.ppf(0.975, df=n - 2) * se

            for scenario, adj in [
                ("Baseline",    0.0),
                ("Optimistic",  -abs(slope_per_quarter) * i * 0.2),
                ("Pessimistic",  abs(slope_per_quarter) * i * 0.3),
            ]:
                val = float(np.clip(projected + adj, 0, 100))
                m   = margin * (0.7 if scenario == "Optimistic" else
                                1.3 if scenario == "Pessimistic" else 1.0)
                projection_records.append({
                    "province":              province,
                    "date":                  future_date,
                    "year":                  future_date.year,
                    "quarter":               future_date.quarter,
                    "quarter_label":         f"Q{future_date.quarter} {future_date.year}",
                    "projected_rate":        round(val, 2),
                    "projected_rate_upper":  round(float(np.clip(val + m, 0, 100)), 2),
                    "projected_rate_lower":  round(float(np.clip(val - m, 0, 100)), 2),
                    "is_projection":         True,
                    "scenario":              scenario,
                })

        proj_2yr = float(np.clip(
            model.predict([[last_numeric + 91 * 8]])[0], 0, 100
        ))
        direction = (
            "Improving" if slope_per_quarter < -0.3 else
            "Worsening" if slope_per_quarter >  0.3 else
            "Stable"
        )
        summary_records.append({
            "province":              province,
            "current_rate":          round(last_rate, 1),
            "projected_rate_2yrs":   round(proj_2yr, 1),
            "change_in_2yrs":        round(proj_2yr - last_rate, 1),
            "trend_direction":       direction,
            "slope_per_quarter_pp":  round(slope_per_quarter, 3),
            "r_squared":             round(r_squared, 3),
            "data_points_used":      n,
        })

    return pd.DataFrame(projection_records), pd.DataFrame(summary_records)

def export_all(df_labour, df_youth_15_34, df_industry, df_gdp,
               df_combined_annual, df_gender, df_neet,
               df_national_sectors, df_projections, df_trend_summary):

    files = {
        "fact_labour_status.csv":          df_labour,
        "fact_youth_national.csv":         df_youth_15_34,
        "fact_industry_employment.csv":    df_industry,
        "fact_provincial_gdp.csv":         df_gdp,
        "fact_unemployment_gdp_annual.csv":df_combined_annual,
        "fact_gender_unemployment.csv":    df_gender,
        "fact_neet_by_province.csv":       df_neet,
        "fact_national_sector_totals.csv": df_national_sectors,
        "fact_projections.csv":            df_projections,
        "fact_trend_summary.csv":          df_trend_summary,
    }
    for fname, df in files.items():
        df.to_csv(CLEAN_DIR + fname, index=False)

    print("\nAll files exported to", CLEAN_DIR)
    for fname in sorted(files):
        rows = sum(1 for _ in open(CLEAN_DIR + fname)) - 1
        print(f"  {fname}: {rows:,} rows")

print("Parsing Table 2.3 (province labour status)...")
df_labour = parse_table_2_3(QLFS_FILE)
df_labour = calculate_derived_metrics(df_labour)
print(f"  {len(df_labour)} rows")

print("Parsing Table 2.2 (age groups / youth)...")
df_age = parse_table_2_2(QLFS_FILE)
df_youth_15_34 = build_youth_15_34(df_age)
print(f"  {len(df_youth_15_34)} rows")

print("Parsing Table 3.2 (industry by province)...")
df_industry = parse_table_3_2(QLFS_FILE)
print(f"  {len(df_industry)} rows")

print("Parsing Table 21 (provincial GDP)...")
df_gdp = parse_table_21(GDP_FILE)
print(f"  {len(df_gdp)} rows")

print("Validating GeoJSON...")
geo_data = load_geojson_for_powerbi(GEOJSON_FILE)

print("Building annual unemployment + GDP join...")
df_labour_annual   = aggregate_to_annual(df_labour)
df_combined_annual = pd.merge(df_labour_annual, df_gdp,
                              on=["province", "year"], how="inner")
df_combined_annual["gdp_per_employed_person"] = (
    df_combined_annual["constant_price_rand"] * 1_000_000 /
    df_combined_annual["employed"]
) / 1000
print(f"  {len(df_combined_annual)} rows")

print("Parsing Table 2 (gender unemployment)...")
df_gender = parse_gender_unemployment(QLFS_FILE)
print(f"  {len(df_gender)} rows")

print("Parsing Table 7c (NEET by province)...")
df_neet = parse_neet_by_province(QLFS_FILE)
print(f"  {len(df_neet)} rows")

print("Parsing Table 3.1 (national sector totals)...")
df_national_sectors = parse_national_sector_totals(QLFS_FILE)
print(f"  {len(df_national_sectors)} rows")

print("Building projections (linear regression per province)...")
df_projections, df_trend_summary = build_projections(df_labour)
print(f"  {len(df_projections)} projection rows, {len(df_trend_summary)} summary rows")

export_all(
    df_labour, df_youth_15_34, df_industry, df_gdp, df_combined_annual,
    df_gender, df_neet, df_national_sectors, df_projections, df_trend_summary
)
